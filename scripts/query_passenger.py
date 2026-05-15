#!/usr/bin/env python3
"""
客流数据查询脚本 v2
从Excel提取客流数据

数据文件：
- 今日数据：~/Desktop/2026年电影小镇实际客流.xlsx

Excel结构（横向表）：
| 门票 | 日期 | 2026-01-01 | 2026-01-02 | ...
| 闸机入园人次 | (第11行) | 11605 | 12832 | ...

第1列: 大类（市场、线上、窗口、闸机入园人次等）
第2列: 子类（研学、散客等）
第1行: 日期
"""

import openpyxl
import json
import datetime
from pathlib import Path

TODAY_FILE = Path.home() / "Desktop/2026年电影小镇实际客流.xlsx"

def find_target_row(ws, target_name):
    """查找指定名称的行"""
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and target_name in str(val):
            return row
    return None

def query_today_data():
    """查询今日客流数据（闸机入园人次）"""
    if not TODAY_FILE.exists():
        return {"error": f"文件不存在: {TODAY_FILE}"}
    
    try:
        wb = openpyxl.load_workbook(str(TODAY_FILE), data_only=True)
        ws = wb.active
        
        # 找到闸机入园人次行
        target_row = find_target_row(ws, "闸机入园人次")
        if not target_row:
            return {"error": "未找到'闸机入园人次'行"}
        
        print(f"找到'闸机入园人次'在第{target_row}行")
        
        # 找最后一列有数据的
        last_val = None
        last_col = None
        last_date = None
        
        for col in range(ws.max_column, 2, -1):
            val = ws.cell(row=target_row, column=col).value
            if val is not None and isinstance(val, (int, float)):
                date_val = ws.cell(row=1, column=col).value
                if isinstance(date_val, datetime.datetime):
                    last_val = val
                    last_col = col
                    last_date = date_val
                    break
        
        if last_val is None:
            return {"error": "没有有效的客流数据"}
        
        return {
            "date": last_date.strftime("%Y-%m-%d"),
            "pax_count": last_val,
            "metric": "闸机入园人次"
        }
        
    except Exception as e:
        return {"error": str(e)}

def query_recent_days(days=7):
    """查询最近N天数据"""
    if not TODAY_FILE.exists():
        return {"error": f"文件不存在: {TODAY_FILE}"}
    
    try:
        wb = openpyxl.load_workbook(str(TODAY_FILE), data_only=True)
        ws = wb.active
        
        target_row = find_target_row(ws, "闸机入园人次")
        if not target_row:
            return {"error": "未找到'闸机入园人次'行"}
        
        # 收集所有有数据的日期
        valid_data = []
        for col in range(3, ws.max_column + 1):
            date_val = ws.cell(row=1, column=col).value
            pax = ws.cell(row=target_row, column=col).value
            if isinstance(date_val, datetime.datetime) and pax is not None:
                valid_data.append((date_val, pax))
        
        # 返回最近days条
        recent = valid_data[-days:] if len(valid_data) >= days else valid_data
        
        return {
            "data": [{"date": d.strftime("%Y-%m-%d"), "pax_count": int(p)} for d, p in recent],
            "count": len(recent)
        }
        
    except Exception as e:
        return {"error": str(e)}

def query_by_date(date_str):
    """查询指定日期的数据"""
    if not TODAY_FILE.exists():
        return {"error": f"文件不存在: {TODAY_FILE}"}
    
    try:
        target_date = datetime.datetime.strptime(date_str, "%Y-%m-%d")
        wb = openpyxl.load_workbook(str(TODAY_FILE), data_only=True)
        ws = wb.active
        
        target_row = find_target_row(ws, "闸机入园人次")
        if not target_row:
            return {"error": "未找到'闸机入园人次'行"}
        
        for col in range(3, ws.max_column + 1):
            date_val = ws.cell(row=1, column=col).value
            if isinstance(date_val, datetime.datetime) and date_val.date() == target_date.date():
                pax = ws.cell(row=target_row, column=col).value
                return {
                    "date": date_str,
                    "pax_count": int(pax) if pax is not None else None,
                    "metric": "闸机入园人次"
                }
        
        return {"error": f"未找到日期: {date_str}"}
        
    except Exception as e:
        return {"error": str(e)}

def validate_data(data):
    """验证数据合理性"""
    issues = []
    
    if "error" in data:
        return {"valid": False, "issues": [data["error"]]}
    
    if "date" not in data or data["date"] is None:
        issues.append("缺少日期")
    
    pax = data.get("pax_count")
    if pax is None:
        issues.append("缺少客流数据")
    elif not isinstance(pax, (int, float)):
        issues.append(f"客流不是数字: {pax}")
    elif pax < 0:
        issues.append(f"客流为负数: {pax}")
    elif pax > 50000:
        issues.append(f"客流异常高: {pax}")
    
    return {"valid": len(issues) == 0, "issues": issues}

if __name__ == "__main__":
    import sys
    
    print("=" * 50)
    print("客流数据查询")
    print("=" * 50)
    
    if len(sys.argv) > 1:
        cmd = sys.argv[1]
        
        if cmd == "today":
            data = query_today_data()
            print(f"\n今日数据: {json.dumps(data, ensure_ascii=False, indent=2)}")
            validation = validate_data(data)
            print(f"\n验证结果: {validation}")
            
        elif cmd == "recent":
            days = int(sys.argv[2]) if len(sys.argv) > 2 else 7
            data = query_recent_days(days)
            print(f"\n最近{days}天数据: {json.dumps(data, ensure_ascii=False, indent=2)}")
            
        elif cmd == "date":
            if len(sys.argv) > 2:
                date_str = sys.argv[2]
                data = query_by_date(date_str)
                print(f"\n{date_str}数据: {json.dumps(data, ensure_ascii=False, indent=2)}")
            else:
                print("用法: python query_passenger.py date <YYYY-MM-DD>")
        else:
            print(f"未知命令: {cmd}")
    else:
        data = query_today_data()
        print(f"\n今日数据: {json.dumps(data, ensure_ascii=False, indent=2)}")
        validation = validate_data(data)
        print(f"\n验证结果: {validation}")
